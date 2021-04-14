import numpy as np
import openpyxl
import string
import datetime
from scipy import interpolate
import calendar
import csv
import os

def write_to_csv(idx,idx1,val,val1,sub_datenum,name):
    daily_values_data=list(map(list,list(zip(*val))))
    daily_values_data1=list(map(list,list(zip(*val1))))

    daily_date_index=list(idx)
    daily_date_index[:] = [x - sub_datenum for x in idx]

    daily_date_index1=list(idx1)
    daily_date_index1[:] = [x - sub_datenum for x in idx1]
    
    for i in range(len(daily_values_data)):
        daily_values_data[i].insert(0,daily_date_index[i])
    for i in range(len(daily_values_data1)):
        daily_values_data1[i].insert(0,daily_date_index1[i])
        
    with open(name+'.csv', 'w') as csvfile:
        spamwriter = csv.writer(csvfile, delimiter=',',quotechar='|', quoting=csv.QUOTE_NONNUMERIC)
        spamwriter.writerows(daily_values_data1)
        spamwriter.writerows(daily_values_data)
        
def filter_data_index(a,thr,method):
    if method =='ge':
        index=list([idx for idx,i in enumerate(a) if i<thr])
    elif method =='le':
        index=list([idx for idx,i in enumerate(a) if i>thr])
    elif method =='l':
        index=list([idx for idx,i in enumerate(a) if i>=thr])
    elif method =='g':
        index=list([idx for idx,i in enumerate(a) if i<=thr])
    return index

def filter_data(a,index,dim):
    b=list(a)
    if dim > 1:
        for i in range(dim):
            b[i] = list(a[i])
            for idx in reversed(index):
                del b[i][idx]
    else:
        for idx in reversed(index):
            del b[idx]
    return b
    
list1=list(string.ascii_uppercase)
list2=['A'+x for x in list1]

list1 = list1+list2

datelist = list1[0:36:3]
valuelist = list1[1:36:3]

Country='UK'# US, EA, JP, UK. 
CurveType='OIS'
PathName=os.getcwd()
ExcelName=os.path.join(PathName, 'A_'+Country+'_All_Data_Bloomberg.xlsx')
if CurveType =='OIS':
    ExcelSheetName='D. Live OIS data'
if CurveType =='Govt':
    ExcelSheetName='D. Live Govt data';

Maturities=[0.25,0.5,1,2,3,4,5,7,10,15,20,30]

wb=openpyxl.load_workbook(ExcelName)
sheet_names=wb.get_sheet_names
sheet=wb['D. Live OIS data']
datenum = [[] for i in range(len(datelist))]
values_data=[[] for i in range(len(datelist))]
for col_num in range(len(datelist)):
    DateGen=sheet[datelist[col_num]+'8':datelist[col_num]+'8000']
    for i in DateGen:
        if i[0].value is not None:
            a=i[0].value.date().toordinal()
            datenum[col_num].append(a)
    ValueGen=sheet[valuelist[col_num]+'8':valuelist[col_num]+'8000']
    
    for i in ValueGen:
        if i[0].value is not None:
            a=i[0].value
            values_data[col_num].append(a)


common_datenum=list(datenum[0])
for i in range(1,len(datenum)):
    common_datenum = list(set(common_datenum)&set(datenum[i]))

common_datenum.sort()
for i in range(len(datenum)):
    diff_dates=list(set(common_datenum) ^ set(datenum[i]))
    for j in diff_dates:
        index=datenum[i].index(j)
        print((str(j)+' date num at '+str(index)+'removed')) 
        value_to_remove = values_data[i][index]
        values_data[i].remove(value_to_remove)

ref_friday=datetime.date(2013,0o5,24).toordinal()
weeks_to_step_back = (ref_friday - common_datenum[0])/7
first_friday = ref_friday - (weeks_to_step_back*7)

week_date_index = list(range(int(first_friday),common_datenum[-1],7))

week_values_data=[[] for i in range(len(datelist))]
for i in range(len(datelist)):
    f=interpolate.interp1d(common_datenum,values_data[i],'zero',bounds_error=False)
    week_values_data[i] = list(f(week_date_index))

month_date_index=[]
first_date=datetime.date.fromordinal(common_datenum[0])
last_date=datetime.date.fromordinal(common_datenum[-1])

first_year = first_date.year
last_year = last_date.year

for i in range(first_date.month,13,1):
    (first_week_day,last_day_of_month)=calendar.monthrange(first_year,i)
    month_date_index.append(datetime.date(first_year,i,last_day_of_month).toordinal())
for i in range(first_year+1,last_year,1):
    for month in range(1,13,1):
        (first_week_day,last_day_of_month)=calendar.monthrange(i,month)
        month_date_index.append(datetime.date(i,month,last_day_of_month).toordinal())
for i in range(1,last_date.month+1,1):
    (first_week_day,last_day_of_month)=calendar.monthrange(last_year,i)
    month_date_index.append(datetime.date(last_year,i,last_day_of_month).toordinal())

month_values_data=[[] for i in range(len(datelist))]
for i in range(len(datelist)):
    f=interpolate.interp1d(common_datenum,values_data[i],'zero',bounds_error=False)
    month_values_data[i] = list(f(month_date_index))

BusinessDayTimestamp = float(common_datenum[-1]-common_datenum[0]+1)/(len(common_datenum)*365.25)
    

#GOVT
if Country =='EA':
    wbG=openpyxl.load_workbook(os.path.join(PathName, 'A_GE_All_Data_Bloomberg.xlsx'))
    sheet=wbG.get_sheet_by_name('D. Live Govt data')
    datenumG = [[] for i in range(len(datelist))]
    values_dataG=[[] for i in range(len(datelist))]
    for col_num in range(len(datelist)):
        DateGen=sheet[datelist[col_num]+'8':datelist[col_num]+'8000']
        for i in DateGen:
            if i[0].value is not None:
                a=i[0].value.date().toordinal()
                datenumG[col_num].append(a)

        ValueGen=sheet[valuelist[col_num]+'8':valuelist[col_num]+'8000']
        
        for i in ValueGen:
            if i[0].value is not None:
                a=i[0].value
                values_dataG[col_num].append(a)

    common_datenumG=list(datenumG[0])
    for i in range(1,len(datenumG)):
        common_datenumG = list(set(common_datenumG)&set(datenumG[i]))

    common_datenumG.sort()

    for i in range(len(datenumG)):
        diff_dates=list(set(common_datenumG) ^ set(datenumG[i]))
        for j in diff_dates:
            index=datenumG[i].index(j)
            print((str(j)+' date num at '+str(index)+'removed')) 
            value_to_remove = values_dataG[i][index]
            values_dataG[i].remove(value_to_remove)

    
    euto_date_num = datetime.date(1999,1,1).toordinal()
    index_to_del=filter_data_index(common_datenumG,euto_date_num,'l')
    pre_euro_ge_index=filter_data(common_datenumG,index_to_del,1)
    pre_euro_ge_vlues=filter_data(values_dataG,index_to_del,len(values_dataG))
    
    index_to_del=filter_data_index(common_datenumG,euto_date_num,'ge')
    post_euro_ge_index=filter_data(common_datenumG,index_to_del,1)
    post_euro_ge_vlues=filter_data(values_dataG,index_to_del,len(values_dataG))

    wbF=openpyxl.load_workbook(os.path.join(PathName, 'A_FR_All_Data_Bloomberg.xlsx'))
    sheet=wbG.get_sheet_by_name('D. Live Govt data')
    datenumF = [[] for i in range(len(datelist))]
    values_dataF=[[] for i in range(len(datelist))]
    for col_num in range(len(datelist)):
        DateGen=sheet[datelist[col_num]+'8':datelist[col_num]+'8000']
        for i in DateGen:
            if i[0].value is not None:
                a=i[0].value.date().toordinal()
                datenumF[col_num].append(a)

        ValueGen=sheet[valuelist[col_num]+'8':valuelist[col_num]+'8000']
        
        for i in ValueGen:
            if i[0].value is not None:
                a=i[0].value
                values_dataF[col_num].append(a)

    common_datenumF=list(datenumF[0])
    for i in range(1,len(datenumF)):
        common_datenumF = list(set(common_datenumF)&set(datenumF[i]))

    common_datenumF.sort()

    for i in range(len(datenumF)):
        diff_dates=list(set(common_datenumF) ^ set(datenumF[i]))
        for j in diff_dates:
            index=datenumG[i].index(j)
            print((str(j)+' date num at '+str(index)+'removed')) 
            value_to_remove = values_dataF[i][index]
            values_dataF[i].remove(value_to_remove)

    
    euto_date_num = datetime.date(1999,1,1).toordinal()

    index_to_del=filter_data_index(common_datenumF,euto_date_num,'ge')
    post_euro_fr_index=filter_data(common_datenumF,index_to_del,1)
    post_euro_fr_vlues=filter_data(values_dataF,index_to_del,len(values_dataF))

    post_inter_index=list(set(post_euro_ge_index)&set(post_euro_fr_index))
    post_inter_index.sort()
    for i in range(len(post_euro_fr_vlues)):
        diff_dates=list(set(post_inter_index) ^ set(post_euro_fr_index))
        for j in diff_dates:
            index=post_euro_fr_index.index(j)
            print((str(j)+' date num at '+str(index)+'removed')) 
            value_to_remove = post_euro_fr_vlues[i][index]
            post_euro_fr_vlues[i].remove(value_to_remove)
    for i in range(len(post_euro_ge_vlues)):
        diff_dates=list(set(post_inter_index) ^ set(post_euro_ge_index))
        for j in diff_dates:
            index=post_euro_ge_index.index(j)
            print((str(j)+' date num at '+str(index)+'removed')) 
            value_to_remove = post_euro_ge_vlues[i][index]
            post_euro_ge_vlues[i].remove(value_to_remove)
    
    common_datenum1 = pre_euro_ge_index + post_inter_index
    values_data1=[[] for i in range(len(post_euro_ge_vlues))]
    for i in range(len(post_euro_ge_vlues)):
        post_euro_sum_values = [x+y for x,y in zip(post_euro_fr_vlues[i],post_euro_ge_vlues[i])]
        post_euro_avg_values=[float(x)*0.5 for x in post_euro_sum_values]
        values_data1[i] = pre_euro_ge_vlues[i] + post_euro_avg_values
    
else:
    sheet=wb['D. Live Govt data']
    datenum1 = [[] for i in range(len(datelist))]
    values_data1=[[] for i in range(len(datelist))]
    for col_num in range(len(datelist)):
        DateGen=sheet[datelist[col_num]+'8':datelist[col_num]+'8000']
        for i in DateGen:
            if i[0].value is not None:
                a=i[0].value.date().toordinal()
                datenum1[col_num].append(a)

        ValueGen=sheet[valuelist[col_num]+'8':valuelist[col_num]+'8000']
        
        for i in ValueGen:
            if i[0].value is not None:
                a=i[0].value
                values_data1[col_num].append(a)

    common_datenum1=list(datenum1[0])
    for i in range(1,len(datenum1)):
        common_datenum1 = list(set(common_datenum1)&set(datenum1[i]))

    common_datenum1.sort()

    for i in range(len(datenum1)):
        diff_dates=list(set(common_datenum1) ^ set(datenum1[i]))
        for j in diff_dates:
            index=datenum1[i].index(j)
            print((str(j)+' date num at '+str(index)+'removed')) 
            value_to_remove = values_data1[i][index]
            values_data1[i].remove(value_to_remove)

ref_friday=datetime.date(2013,0o5,24).toordinal()
weeks_to_step_back = (ref_friday - common_datenum1[0])/7
first_friday = ref_friday - (weeks_to_step_back*7)

week_date_index1 = list(range(int(first_friday),common_datenum1[-1],7))
week_values_data1=[[] for i in range(len(datelist))]
for i in range(len(datelist)):
    f=interpolate.interp1d(common_datenum1,values_data1[i],'zero',bounds_error=False)
    week_values_data1[i] = list(f(week_date_index1))

month_date_index1=[]
first_date=datetime.date.fromordinal(common_datenum1[0])
last_date=datetime.date.fromordinal(common_datenum1[-1])
first_year = first_date.year
last_year = last_date.year

for i in range(first_date.month,13,1):
    (first_week_day,last_day_of_month)=calendar.monthrange(first_year,i)
    month_date_index1.append(datetime.date(first_year,i,last_day_of_month).toordinal())
for i in range(first_year+1,last_year,1):
    for month in range(1,13,1):
        (first_week_day,last_day_of_month)=calendar.monthrange(i,month)
        month_date_index1.append(datetime.date(i,month,last_day_of_month).toordinal())
for i in range(1,last_date.month+1,1):
    (first_week_day,last_day_of_month)=calendar.monthrange(last_year,i)
    month_date_index1.append(datetime.date(last_year,i,last_day_of_month).toordinal())

month_values_data1=[[] for i in range(len(datelist))]
for i in range(len(datelist)):
    f=interpolate.interp1d(common_datenum1,values_data1[i],'zero',bounds_error=False)
    month_values_data1[i] = list(f(month_date_index1))

BusinessDayTimestamp1 = float(common_datenum1[-1]-common_datenum1[0]+1)/(len(common_datenum1)*365.25)

sub_datenum = datetime.date(1899,12,30).toordinal()

if Country == 'JP':
    firstday=datetime.date(2009,8,6).toordinal()
else:
    firstday=common_datenum[0]

lastday=firstday-1


index_to_del=filter_data_index(common_datenum,firstday,'ge')
new_daily_index=filter_data(common_datenum,index_to_del,1)
new_daily_values=filter_data(values_data,index_to_del,len(values_data))

index_to_del=filter_data_index(common_datenum1,lastday,'le')
new_daily_index1=filter_data(common_datenum1,index_to_del,1)
new_daily_values1=filter_data(values_data1,index_to_del,len(values_data1))

write_to_csv(new_daily_index,new_daily_index1,new_daily_values,new_daily_values1,sub_datenum,Country+'_Daily')

index_to_del=filter_data_index(week_date_index,firstday,'ge')
new_daily_index=filter_data(week_date_index,index_to_del,1)
new_daily_values=filter_data(week_values_data,index_to_del,len(week_values_data))

index_to_del=filter_data_index(week_date_index1,lastday,'le')
new_daily_index1=filter_data(week_date_index1,index_to_del,1)
new_daily_values1=filter_data(week_values_data1,index_to_del,len(week_values_data1))
write_to_csv(new_daily_index,new_daily_index1,new_daily_values,new_daily_values1,sub_datenum,Country+'_Weekly')

index_to_del=filter_data_index(month_date_index,firstday,'ge')
new_daily_index=filter_data(month_date_index,index_to_del,1)
new_daily_values=filter_data(month_values_data,index_to_del,len(month_values_data))

index_to_del=filter_data_index(month_date_index1,lastday,'le')
new_daily_index1=filter_data(month_date_index1,index_to_del,1)
new_daily_values1=filter_data(month_values_data1,index_to_del,len(month_values_data1))

write_to_csv(new_daily_index,new_daily_index1,new_daily_values,new_daily_values1,sub_datenum,Country+'_Monthly')


